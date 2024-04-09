# Создаёт Excel нужной структуры
import openpyxl
from openpyxl.utils import get_column_letter

def create_excel_file():
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()

    # Получаем активный лист и переименовываем его в 'Rent'
    rent_sheet = wb.active
    rent_sheet.title = 'Rent'

    # Добавляем заголовки для листа аренды
    rent_headers = ['Date', 'Tenant', 'Type', 'Amount', 'Description']
    for col, header in enumerate(rent_headers, start=1):
        rent_sheet[get_column_letter(col) + '1'] = header

    # Создаем лист для заметок
    notes_sheet = wb.create_sheet(title='Notes')

    # Добавляем заголовки для листа заметок
    notes_headers = ['Date', 'Note']
    for col, header in enumerate(notes_headers, start=1):
        notes_sheet[get_column_letter(col) + '1'] = header

    # Сохраняем книгу
    wb.save('finances.xlsx')

if __name__ == '__main__':
    create_excel_file()
